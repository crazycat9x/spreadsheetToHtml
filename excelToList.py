import os
import odfToArray

def excelToList(target):
    file_ext = target[target.index(".")+1:]
    data = []
    if file_ext == "xls" or file_ext == "xlsx":
        import xlrd
        wb = xlrd.open_workbook(target)
        sh1 = wb.sheet_by_index(0)
        for rownum in range(sh1.nrows): 
            data += [sh1.row_values(rownum)]
    elif file_ext == "csv":
        import csv
        reader = csv.reader(open(target, "rb"))
        for row in reader:
            data += [row]
    elif file_ext == "lsx":
        from openpyxl.reader.excel import load_workbook
        wb = load_workbook(target=target, use_iterators = True)
        sheet = wb.get_active_sheet()
        for row in sheet.iter_rows():
            data_row = []
            for cell in row:
                data_row += [cell.internal_value]
            data += [data_row]
    elif file_ext == "ods":
        from odsreader import ODSReader
        doc = ODSReader(target)
        table = doc.SHEETS.items()[0]
        data += table[1]
    return data
